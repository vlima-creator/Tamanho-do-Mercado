#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para análise detalhada do Template de Análise de Mercado Marketplace
"""

import openpyxl
from openpyxl import load_workbook
import pandas as pd
import sys
import json

def analyze_excel_structure(file_path):
    """Analisa a estrutura completa do arquivo Excel"""
    
    print("="*80)
    print("ANÁLISE DO TEMPLATE DE ANÁLISE DE MERCADO MARKETPLACE")
    print("="*80)
    print()
    
    # Carregar o workbook
    wb = load_workbook(file_path, data_only=False)
    
    print(f"📊 Total de planilhas: {len(wb.sheetnames)}")
    print()
    
    analysis_results = {}
    
    for sheet_name in wb.sheetnames:
        print("="*80)
        print(f"📄 PLANILHA: {sheet_name}")
        print("="*80)
        
        ws = wb[sheet_name]
        
        # Informações básicas
        print(f"\n📐 Dimensões:")
        print(f"   - Linhas com dados: {ws.max_row}")
        print(f"   - Colunas com dados: {ws.max_column}")
        
        # Tentar ler com pandas para análise mais detalhada
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
            
            print(f"\n📋 Conteúdo da planilha:")
            print(f"   - Dimensões do DataFrame: {df.shape}")
            
            # Mostrar as primeiras linhas não vazias
            print(f"\n🔍 Primeiras linhas com conteúdo:")
            non_empty_rows = df.dropna(how='all').head(15)
            
            for idx, row in non_empty_rows.iterrows():
                row_data = row.dropna()
                if not row_data.empty:
                    print(f"\n   Linha {idx + 1}:")
                    for col_idx, value in row_data.items():
                        if pd.notna(value) and str(value).strip() != '':
                            print(f"      Col {col_idx + 1}: {str(value)[:100]}")
            
            # Identificar cabeçalhos potenciais
            print(f"\n📌 Identificação de estrutura:")
            
            # Procurar por células mescladas
            if hasattr(ws, 'merged_cells'):
                merged_ranges = list(ws.merged_cells.ranges)
                if merged_ranges:
                    print(f"\n   ⚠️  Células mescladas encontradas: {len(merged_ranges)}")
                    for merge_range in merged_ranges[:5]:  # Mostrar primeiras 5
                        print(f"      - {merge_range}")
            
            # Procurar por padrões de dados
            print(f"\n   🔢 Análise de tipos de dados:")
            for col in range(min(10, df.shape[1])):  # Analisar primeiras 10 colunas
                col_data = df[col].dropna()
                if len(col_data) > 0:
                    # Identificar tipo predominante
                    numeric_count = col_data.apply(lambda x: isinstance(x, (int, float))).sum()
                    string_count = col_data.apply(lambda x: isinstance(x, str)).sum()
                    
                    if numeric_count > string_count:
                        print(f"      Coluna {col + 1}: Predominantemente NUMÉRICA ({numeric_count} valores)")
                    elif string_count > 0:
                        print(f"      Coluna {col + 1}: Predominantemente TEXTO ({string_count} valores)")
            
            # Salvar dados para análise posterior
            analysis_results[sheet_name] = {
                'dimensions': df.shape,
                'sample_data': df.head(20).to_dict(),
                'has_data': not df.dropna(how='all').empty
            }
            
        except Exception as e:
            print(f"\n   ⚠️  Erro ao ler com pandas: {str(e)}")
            
            # Análise alternativa célula por célula
            print(f"\n   📝 Análise célula por célula (primeiras 20 linhas):")
            for row_idx in range(1, min(21, ws.max_row + 1)):
                row_content = []
                for col_idx in range(1, min(11, ws.max_column + 1)):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.value is not None:
                        row_content.append(f"Col{col_idx}: {str(cell.value)[:50]}")
                
                if row_content:
                    print(f"\n      Linha {row_idx}: {' | '.join(row_content)}")
        
        print()
    
    # Resumo final
    print("\n" + "="*80)
    print("📊 RESUMO DA ANÁLISE")
    print("="*80)
    print(f"\nTotal de planilhas analisadas: {len(analysis_results)}")
    print("\nPlanilhas com dados:")
    for sheet_name, data in analysis_results.items():
        if data['has_data']:
            print(f"   ✓ {sheet_name}: {data['dimensions'][0]} linhas x {data['dimensions'][1]} colunas")
        else:
            print(f"   ✗ {sheet_name}: (vazia)")
    
    print("\n" + "="*80)

if __name__ == "__main__":
    file_path = "/home/user/uploaded_files/Template_Analise_Mercado_Marketplace_v8_Sheets (1).xlsx"
    analyze_excel_structure(file_path)
